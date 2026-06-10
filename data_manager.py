import pandas as pd
from openpyxl import load_workbook

PATH_EXCEL = "caja.xlsx"

def guardar_registro(inversion, fecha, nominales, saldo):
    """Inserta una nueva fila guardando la fecha con formato nativo de Excel (dd/mm/yyyy)."""
    try:
        wb = load_workbook(PATH_EXCEL)
        if inversion in wb.sheetnames:
            ws = wb[inversion]
            
            # 1. Buscamos la verdadera última fila con datos
            verdadera_ultima_fila = ws.max_row
            while verdadera_ultima_fila > 0:
                valores_linea = [
                    ws.cell(row=verdadera_ultima_fila, column=1).value,
                    ws.cell(row=verdadera_ultima_fila, column=2).value,
                    ws.cell(row=verdadera_ultima_fila, column=3).value
                ]
                if any(v is not None and str(v).strip() != "" for v in valores_linea):
                    break
                verdadera_ultima_fila -= 1
            
            fila_destino = verdadera_ultima_fila + 1
            
            # 2. Escribimos la FECHA como objeto nativo y le aplicamos el formato numérico de Excel
            # Nota técnica: Aunque en tu Excel veas "aaaa", internamente openpyxl usa "yyyy" por el estándar global.
            celda_fecha = ws.cell(row=fila_destino, column=1, value=fecha)
            celda_fecha.number_format = 'dd/mm/yyyy'
            
            # 3. Escribimos el resto de los datos
            ws.cell(row=fila_destino, column=2, value=nominales)
            ws.cell(row=fila_destino, column=3, value=saldo)
            
            wb.save(PATH_EXCEL)
            return True, f"¡Guardado con éxito en '{inversion}' (Fila {fila_destino})!"
            
        return False, f"La hoja '{inversion}' no existe."
    except Exception as e:
        return False, f"Error al escribir: {e}. ¿Está abierto el Excel?"

def cargar_todo():
    """Lee todas las hojas, las unifica y asegura el tipo de datos numérico."""
    try:
        hojas = pd.read_excel(PATH_EXCEL, sheet_name=None, engine="openpyxl")
        lista_df = []
        for nombre, df in hojas.items():
            df.columns = df.columns.str.strip()
            if all(c in df.columns for c in ['Fecha actualizacion', 'nominales', 'saldo']):
                df = df.rename(columns={'Fecha actualizacion': 'Fecha', 'nominales': 'Nominales', 'saldo': 'Saldo'})
                
                # Normalizar fechas y forzar números (Clave para que Plotly no falle)
                df['Fecha'] = pd.to_datetime(df['Fecha']).dt.normalize()
                df['Saldo'] = pd.to_numeric(df['Saldo'], errors='coerce')
                df['Nominales'] = pd.to_numeric(df['Nominales'], errors='coerce')
                
                df['Inversion'] = nombre
                
                # Descartar filas vacías de esta hoja
                df = df.dropna(subset=['Fecha', 'Saldo'])
                lista_df.append(df)
                
        return pd.concat(lista_df, ignore_index=True) if lista_df else pd.DataFrame()
    except:
        return pd.DataFrame()